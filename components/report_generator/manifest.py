import openpyxl
import pandas as pd

class Manifest:
    """
    A class to represent a manifest file for the report generator.
    The manifest file contains a timeline of the report generation process as well as the list of students and the comments generated for each student.
    The manifest file is in Excel format.
    """

    def __init__(self, file_path):
        """
        Constructs a new instance of the manifest file.

        Args:
            file_path: The file path of the manifest file.
        """
        self.file_path = file_path

        # Create new pandas dataframe to store the manifest data
        self.table = pd.DataFrame(columns = ["Student", 
                                             "Original Comment", 
                                             "Shortened Comment", 
                                             "Final Comment Length (Chars)", 
                                             "Final Comment Length (Words)", 
                                             "Status", 
                                             "Completed At", 
                                             "Error"])

    def save(self):
        """
        Saves the manifest file.
        """
        self.table.to_excel(self.file_path, sheet_name = "Report Manifest", index = False)

        # Enable word wrap for the "Original Comment" and "Shortened Comment" columns
        workbook = openpyxl.load_workbook(self.file_path)
        worksheet = workbook.active
        worksheet.column_dimensions["A"].width = 15
        worksheet.column_dimensions["B"].width = 75
        worksheet.column_dimensions["C"].width = 75
        worksheet.column_dimensions["D"].width = 25
        worksheet.column_dimensions["E"].width = 25
        worksheet.column_dimensions["F"].width = 10
        worksheet.column_dimensions["G"].width = 25
        worksheet.column_dimensions["H"].width = 40

        for row in worksheet.iter_rows(min_row = 2, max_row = worksheet.max_row, min_col = 2, max_col = 3):
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(wrap_text = True)

        workbook.save(self.file_path)
        workbook.close()

    def add_entry(self, 
                  student,
                  comment_orig,
                  comment_short,
                  length_chars,
                  length_words,
                  status,
                  completed_at,
                  error = None):
        """
        Adds an entry to the manifest file.
        
        Args:
            student: The student's name.
            comment_orig: The original comment.
            comment_short: The shortened comment.
            length_chars: The length of the shortened comment in characters.
            length_words: The length of the shortened comment in words.
            status: The status of the report generation process.
            completed_at: The time when the report generation process was completed.
            error: The error message if the report generation process failed.
        """
        self.table = pd.concat(
            [
                pd.DataFrame(
                    [[
                        student, 
                        comment_orig, 
                        comment_short, 
                        length_chars,
                        length_words,
                        status, 
                        completed_at, 
                        error
                    ]], 
                    columns = self.table.columns
                ),
                self.table
            ], 
            ignore_index = True
        )