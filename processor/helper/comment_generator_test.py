# TEST
import itertools
import getopt
import sys

import pandas as pd
import openpyxl
import language_tool_python as ltp

help_text = """
HELP PAGE
=========
This script generates comments for the given source file, checks the grammar, and saves the results to the given output file.

=========
USAGE
=========
Format:
comment_generator.py -s <source_file_path> -o <output_file_path>

Options:
-h, --help
    Displays this help page.
-s, --source <source_file_path>
    Specifies the path of the source file (grader report). Must be an Excel file (.xlsm or .xlsx).
    Note: The source file must contain a sheet named 'Comment Mapping'.
    Example: 
        C:/Users/John Doe/Desktop/Grader Report P1A Art Sample.xlsm
        or
        C:/Users/John Doe/Desktop/Grader Report P1A Art Sample.xlsx
-o, --output <output_file_path>
    Specifies the path of the output file (Excel file). Must include the file name and extension.
    Note: If the file already exists, it will be overwritten.
    Example: 
        C:/Users/John Doe/Desktop/Test Result.xlsx

Example:
comment_generator.py -s C:/Users/John Doe/Desktop/Grader Report P1A Art Sample.xlsm -o C:/Users/John Doe/Desktop/Test Result.xlsx

Note:
The script will generate comments for all possible combinations of grades for each criteria.
"""

def run(source_file_path, output_file_path, ltm, CommentGenerator):
    tool = ltm.get_tool()
    comment_mapping = pd.read_excel(source_file_path, sheet_name = "Comment Mapping", index_col = 0, header = 0)
    result = pd.DataFrame(columns = ["FG", "Pos", "Neg", "Comment", "Language Issue", "Possible Correction"])

    criteria = comment_mapping.drop(index = ["Intro", "Closing"]).index.tolist()
    print(f"[  ] {len(criteria)} criterias found:\n     {criteria}")

    for i in range(len(criteria)):
        result.insert(i, criteria[i], "")

    grades = ["A", "B", "C", "D"]

    print(f"[  ] Generating comments for {4 * (len(grades) ** len(criteria))} possible combinations of grades…")
    for student_final_grade in grades:
        print(f"[  ] Generating comments for {student_final_grade}…")
        for grade_combination in itertools.product(grades, repeat = len(criteria)):
            student_result = dict(zip(criteria, grade_combination))
            comment, pos, neg = CommentGenerator("Yabushita Fu", "Fu", "F", comment_mapping, student_result, student_final_grade).generate_comment(probe = True, autocorrect = False, ltm = ltm)
            result.loc[len(result)] = list(grade_combination) + [student_final_grade, pos, neg, comment, "", ""]

    # Remove duplicate comments
    print(f"[  ] Removing duplicate comments…")
    result = result.drop_duplicates(subset = ["Comment"])

    # Check grammar of comments
    print(f"[  ] Checking grammar of comments…")
    tool = ltp.LanguageTool('en-UK')
    for index, row in result.iterrows():
        matches = tool.check(row["Comment"])
        for match in matches:
            result.loc[index, "Language Issue"] = f"{result.loc[index, 'Language Issue']}{match.ruleId}: {match.message}; "
            result.loc[index, "Possible Correction"] = f"{tool.correct(row['Comment'])};"

    tool.close()

    # Save results to Excel
    print(f"[**] Saving results to Excel…")
    result.to_excel(output_file_path, sheet_name = "Unique Outputs", index = False)

    # Format Excel sheet
    print(f"[  ] Formatting Excel sheet…")
    wb = openpyxl.load_workbook(output_file_path)
    ws = wb["Unique Outputs"]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    ws.row_dimensions[1].height = 75

    for i in range(1, len(criteria) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 15

    for i in range(len(criteria) + 1, len(criteria) + 4):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 5

    ws.column_dimensions[openpyxl.utils.get_column_letter(len(criteria) + 4)].width = 70
    ws.column_dimensions[openpyxl.utils.get_column_letter(len(criteria) + 5)].width = 40
    ws.column_dimensions[openpyxl.utils.get_column_letter(len(criteria) + 6)].width = 70

    for row in ws.iter_rows(min_col = 1, max_col = len(criteria) + 6, min_row = 1, max_row = 1):
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(wrap_text = True, vertical = "center", horizontal = "center")

    for row in ws.iter_rows(min_col = 1, max_col = len(criteria) + 3, min_row =  2, max_row = len(result) + 1):
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(vertical = "center", horizontal = "center")

    for row in ws.iter_rows(min_col = len(criteria) + 4, max_col = len(criteria) + 6, min_row = 2, max_row = len(result) + 1):
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(wrap_text = True, vertical = "center", horizontal = "left")

    print(f"[**] Saving Excel sheet…")
    wb.save(output_file_path)

    print(f"[  ] Done.")

def main(argv):
    import language_tool_master as ltm
    from comment_generator import CommentGenerator

    source_file_path = ""
    output_file_path = ""

    try:
        opts, args = getopt.getopt(argv, "hs:o:", ["help", "source=", "output="])
    except getopt.GetoptError:
        print("Error! Invalid argument(s).")
        print("comment_generator.py -s <source_file_path> -o <output_file_path> --help")
        sys.exit(2)

    for opt, arg in opts:
        if opt in ("-h", "--help"):
            print(help_text)
            sys.exit()
        elif opt in ("-s", "--source"):
            source_file_path = arg
        elif opt in ("-o", "--output"):
            output_file_path = arg

    if source_file_path == "" or output_file_path == "":
        print("Error! No output and/or source file path specified.")
        print("comment_generator.py -s <source_file_path> -o <output_file_path>")
        sys.exit(2)

    run(source_file_path, output_file_path, ltm, CommentGenerator)

if __name__ == "__main__":
    main(sys.argv[1:])